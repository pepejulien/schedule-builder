#!/usr/bin/env python3
"""
JAJB Driver Schedule BUILDER  (generates a weekly schedule from scratch)
========================================================================
Distinct from `weekly-driver-schedule` (which FORMATS an already-made schedule).
This one ASSIGNS drivers to a new week given route demand + availability.

INPUTS (all via a single --config JSON; see SKILL.md for the schema):
  - avail_file : the week's uploaded "Shifts & Availability" xlsx whose cells
                 carry each driver's submitted `Unavailable` days (HARD constraint).
  - prev_week_file : last week's schedule xlsx, for the consecutive-day carryover.
  - prefs_csv : Driver-Preferences.csv (derived usual/often-off days = SOFT tiebreak;
                and, with merge_standing_unavailable, the STANDING hard days-off).
  - waves : per-operating-day dict of {wave_time: route_count} (EXACT).
  - backups : per-day count, or a percent of routes (Jose's band 10-20%; default 15%).
  - exclude / exact_days / reduced_days / most_days : roster & per-driver day targets.

RULES enforced (the invariants the pytest suite locks in):
  1. HARD: never schedule a driver on a day they marked Unavailable.
  2. HARD: max N consecutive WORKED days (default 5), counting the prior week's
     tail (backup days count as worked, per dispatch policy).
  3. Per-wave route counts hit EXACTLY.
  4. No overtime: a weekly day cap (max_primary_days) and hours cap
     (weekly_hours_cap). Backups are a TOP-UP only -- assigned to regular drivers
     who already have >=2 primary days, never pushing anyone over the hours cap,
     and never as someone's whole week (no backup-only weeks).
  5. Day targets: exact_days (exact count, capped by the weekly cap),
     reduced_days (fewer-than-average target), most_days (maximize up to the cap);
     everyone else balanced to a tight hours band.
  6. Soft preference: usual-days from the prefs CSV are a TIEBREAK only.
  7. Exclusions: drop the dispatch/management names listed in the config.

HARDENING (2026-06):
  * Runs natively on Windows -- the old %-m/%-d/%-I strftime format crashed here.
  * Every string written to the xlsx is ASCII-safe (no mojibake em-dashes).
  * Config is validated on load; bad/missing fields fail loudly with a clear message.
  * A named driver (exclude / target / most) that doesn't match the roster is a
    LOUD error by default (strict_names), not a silent skip.
  * The solver core (build_schedule) is importable and returns a Result object so
    the test suite can assert the invariants directly.

OPTIONAL FEATURES (behind config flags):
  * merge_standing_unavailable (default off) : union each driver's STANDING
    days-off (unavailable_hard in the prefs CSV) into that week's hard
    Unavailable set, so a missed weekly submission can't schedule someone on
    a permanent day off.
  * max_weekend_days (default off) : when both Sat and Sun are operating days,
    HARD-cap each driver's weekend days.
  * use_premade_shifts (default ON) : pre-filled shift cells in avail_file are
    SEED days -- Jose's pre-made schedule, the strongest soft preference.
    Kept whenever the rules allow; times ignored (waves re-assigned freely).
  * weekend_spread (default ON) : soft nudge toward ~1 weekend day per driver
    when Sat+Sun both operate -- not half the fleet working both weekend days
    while the other half has none. Company need still wins.

OUTPUT: the configured xlsx with a `Shifts & Availability` sheet (same format as
the inputs, ready to enter into the system) + a `By Day` sheet, plus a verification
summary to stdout. Exits non-zero if any invariant check fails.

Run:  python build_weekly_schedule.py --config Week-NN-config.json
"""
import argparse, json, re, datetime, sys, os
from collections import Counter
import openpyxl, warnings
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
warnings.filterwarnings('ignore')

ALL = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
WEEKEND = {'Sat', 'Sun'}
SHEET = 'Shifts & Availability'


class ScheduleConfigError(Exception):
    """Raised on any invalid/inconsistent config so the run fails loudly."""
    pass


# ---------------------------------------------------------------- helpers ----
def norm(s):
    return re.sub(r'\s+', ' ', str(s)).strip().lower()


# Map the common non-ASCII characters a schedule might pick up to plain ASCII,
# then hard-strip anything else. Keeps every file write ASCII-safe (no mojibake).
_UNI = {
    '—': '-', '–': '-', '‒': '-', '‑': '-', '‐': '-',
    '‘': "'", '’': "'", '“': '"', '”': '"',
    '…': '...', '•': '*', ' ': ' ', '�': '',
}


def asciize(v):
    """Return v unchanged for non-strings; for strings, an ASCII-only version."""
    if not isinstance(v, str):
        return v
    for k, r in _UNI.items():
        v = v.replace(k, r)
    return v.encode('ascii', 'ignore').decode('ascii')


def fmt_timestamp(dt):
    """'6/27/26, 2:30:04 AM' without platform-specific strftime (%-m crashes on Windows)."""
    h = dt.hour % 12 or 12
    ap = 'AM' if dt.hour < 12 else 'PM'
    return f"{dt.month}/{dt.day}/{dt.strftime('%y')}, {h}:{dt.strftime('%M:%S')} {ap}"


def _open_shifts_sheet(path):
    """Load the 'Shifts & Availability' sheet, tolerating extra leading sheets."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET in wb.sheetnames:
        return wb[SHEET]
    # fall back: first sheet whose header row names an associate column
    for ws in wb.worksheets:
        for r in range(1, min(ws.max_row, 8) + 1):
            if str(ws.cell(r, 1).value or '').strip().lower().startswith('associate'):
                return ws
    raise ScheduleConfigError(
        f"'{SHEET}' sheet not found in {path} (sheets: {wb.sheetnames})")


def _layout(ws):
    """Locate the name / Transporter ID / day columns from the header row
    instead of assuming fixed positions -- the sheet may carry extra columns
    (e.g. the Tier column Jose added 2026-07-11). Returns
    (name_col, tid_col, {day: col}, first_data_row)."""
    for r in range(1, min(ws.max_row, 8) + 1):
        vals = {c: str(ws.cell(r, c).value or '').strip() for c in range(1, ws.max_column + 1)}
        namec = next((c for c, v in vals.items() if v.lower().startswith('associate')), None)
        if not namec:
            continue
        daycols = {}
        for c, v in vals.items():
            m = re.match(r'(Sun|Mon|Tue|Wed|Thu|Fri|Sat)\b', v)
            if m:
                daycols[m.group(1)] = c
        if len(daycols) == 7:
            tidc = next((c for c, v in vals.items() if 'transporter' in v.lower()), namec + 1)
            return namec, tidc, daycols, r + 1
    # legacy fixed layout: names col 1, TID col 2, days cols 3-9, data from row 6
    return 1, 2, {d: 3 + j for j, d in enumerate(ALL)}, 6


def _is_data_name(v):
    s = str(v or '').strip()
    return bool(s) and not s.lower().startswith('total')


def load_roster(avail_file):
    """Roster + hard Unavailable days + SEED days. A seed is any pre-filled
    shift cell (anything non-blank that isn't 'Unavailable') in the uploaded
    file -- Jose's pre-made schedule. Seed days are the strongest SOFT
    preference: kept whenever the rules allow, dropped freely when a tier
    cap / wave count / consecutive rule forces it. Times in seed cells are
    ignored (waves are re-assigned as needed, per Jose 2026-07-04)."""
    ws = _open_shifts_sheet(avail_file)
    namec, tidc, daycols, r0 = _layout(ws)
    rows = []
    for r in range(r0, ws.max_row + 1):
        nm = ws.cell(r, namec).value
        if not _is_data_name(nm):
            continue
        name = re.sub(r'\s+', ' ', str(nm)).strip()
        unav, seed, meet, meet_txt = set(), set(), set(), {}
        for d, col in daycols.items():
            v = ws.cell(r, col).value
            if not v or not str(v).strip():
                continue
            lv = str(v).lower()
            if 'unavail' in lv:
                unav.add(d)
            elif 'meeting' in lv:
                # DO-NOT-TOUCH day (Jose 2026-07-10): a pre-entered ~2h meeting.
                # Preserved verbatim in the output; no route/backup that day;
                # still counts as a worked (2h) day for the consecutive rule
                # and the total-days cap.
                meet.add(d); meet_txt[d] = str(v).strip()
            elif not any(t in lv for t in ('closed', 'dispatch')):
                seed.add(d)
        rows.append(dict(name=name, tid=str(ws.cell(r, tidc).value or '').strip(),
                         unav=unav, seed=seed, meet=meet, meet_txt=meet_txt,
                         std_added=set(), usual=[], soft=[],
                         present=0, prim=[], bk=[], helper=[], extra=set()))
    if not rows:
        raise ScheduleConfigError(f"No driver rows found in {avail_file} (no name rows under the header).")
    return rows


def load_prev_worked(prev_file, start_date):
    """Return {norm_name: set(date)} of days worked the prior week (by column position)."""
    if not prev_file:
        return {}
    ws = _open_shifts_sheet(prev_file)
    namec, _tidc, daycols, r0 = _layout(ws)
    base = start_date - datetime.timedelta(days=7)
    coldate = {ALL[j]: base + datetime.timedelta(days=j) for j in range(7)}
    out = {}
    for r in range(r0, ws.max_row + 1):
        nm = ws.cell(r, namec).value
        if not _is_data_name(nm):
            continue
        s = set()
        for d, col in daycols.items():
            v = ws.cell(r, col).value
            if v and str(v).strip() and 'unavailable' not in str(v).lower():
                s.add(coldate[d])
        out[norm(nm)] = s
    return out


def load_prefs(prefs_csv):
    import csv
    p = {}
    if not prefs_csv:
        return p
    with open(prefs_csv, encoding='utf-8') as f:
        for row in csv.DictReader(f):
            p[norm(row['driver'])] = dict(
                usual=[x for x in row.get('usual_days', '').split('|') if x],
                soft=[x for x in row.get('often_off_soft', '').split('|') if x],
                hard=[x for x in row.get('unavailable_hard', '').split('|') if x],
                present=int(row.get('weeks_present', 0) or 0))
    return p


def resolve(name, roster_names):
    """Fuzzy: every token of the query must be a substring of some token of the
    candidate. An EXACT (normalized) match short-circuits first, so a config
    name that IS a roster name is never reported ambiguous just because it is
    also a prefix of a longer one (e.g. "Robert Smith" vs "Robert Smith Jr")."""
    nq = norm(name)
    if nq in roster_names:
        return [nq]
    q = nq.split()
    return [n for n in roster_names
            if all(any(tok in rt for rt in n.split()) for tok in q)]


# -------------------------------------------------------- config handling ----
def load_config(config_path):
    """Read + validate the config JSON. Resolves relative input/output paths
    against the config file's own directory. Raises ScheduleConfigError loudly."""
    try:
        with open(config_path, encoding='utf-8') as f:
            cfg = json.load(f)
    except FileNotFoundError:
        raise ScheduleConfigError(f"Config file not found: {config_path}")
    except json.JSONDecodeError as e:
        raise ScheduleConfigError(f"Config is not valid JSON ({config_path}): {e}")

    base = os.path.dirname(os.path.abspath(config_path))

    def rp(p):
        if not p:
            return p
        return p if os.path.isabs(p) else os.path.normpath(os.path.join(base, p))

    for k in ('avail_file', 'prev_week_file', 'prefs_csv', 'out'):
        if cfg.get(k):
            cfg[k] = rp(cfg[k])

    errs = []

    # required structure
    if 'start_date' not in cfg:
        errs.append("missing 'start_date' (the Sunday of the week, YYYY-MM-DD)")
    else:
        try:
            d = datetime.date.fromisoformat(cfg['start_date'])
            if d.weekday() != 6:  # Monday=0 .. Sunday=6
                errs.append(f"start_date {cfg['start_date']} is a "
                            f"{d.strftime('%A')}, not a Sunday")
        except ValueError:
            errs.append(f"start_date '{cfg['start_date']}' is not a valid ISO date")

    waves = cfg.get('waves')
    if not isinstance(waves, dict) or not waves:
        errs.append("'waves' must be a non-empty object of {day: {wave_time: count}}")
    else:
        for d, w in waves.items():
            if d not in ALL:
                errs.append(f"waves: '{d}' is not a valid day ({ALL})")
            if not isinstance(w, dict) or not w:
                errs.append(f"waves['{d}'] must be a non-empty object of wave counts")
                continue
            for t, n in w.items():
                if not isinstance(n, int) or n < 0:
                    errs.append(f"waves['{d}']['{t}'] must be a non-negative integer, got {n!r}")

    # required input file must exist (fail loud, not a half-built schedule)
    if not cfg.get('avail_file'):
        errs.append("missing 'avail_file' (this week's availability xlsx)")
    elif not os.path.isfile(cfg['avail_file']):
        errs.append(f"avail_file not found: {cfg['avail_file']}")

    # prev_week_file: required for the consecutive-day carryover unless explicitly null
    if 'prev_week_file' not in cfg:
        errs.append("missing 'prev_week_file' (set to null to disable the consecutive-day "
                    "carryover, but then invariant 2 can't span the week boundary)")
    elif cfg['prev_week_file'] and not os.path.isfile(cfg['prev_week_file']):
        errs.append(f"prev_week_file not found: {cfg['prev_week_file']}")

    if cfg.get('prefs_csv') and not os.path.isfile(cfg['prefs_csv']):
        errs.append(f"prefs_csv not found: {cfg['prefs_csv']}")
    if cfg.get('merge_standing_unavailable') and not cfg.get('prefs_csv'):
        errs.append("merge_standing_unavailable is on but no prefs_csv is set")

    if not cfg.get('out'):
        errs.append("missing 'out' (output xlsx path)")

    # numeric knobs
    for k in ('max_consecutive', 'max_primary_days', 'free_primary_cap',
              'primary_hours', 'backup_hours', 'weekly_hours_cap', 'max_weekend_days',
              'max_total_days', 'free_total_days'):
        if k in cfg and cfg[k] is not None and not isinstance(cfg[k], (int, float)):
            errs.append(f"'{k}' must be a number, got {cfg[k]!r}")

    fb = cfg.get('backup_fallback')
    if fb is not None and (not isinstance(fb, list)
                           or any(not isinstance(g, list) for g in fb)):
        errs.append("'backup_fallback' must be a list of name-lists, ordered by "
                    "priority (e.g. [[top performers], [solid], [underperforming]])")

    dr_rates = cfg.get('driver_rates')
    if dr_rates is not None and (not isinstance(dr_rates, dict)
                                 or any(not isinstance(v, (int, float))
                                        for v in dr_rates.values())):
        errs.append("'driver_rates' must be {name: rate} numbers (board Rate, closer to 0 = better)")

    tp = cfg.get('training_pairs')
    if tp is not None and (not isinstance(tp, list)
                           or any(not isinstance(p, dict) or 'trainer' not in p
                                  or 'trainee' not in p for p in tp)):
        errs.append("'training_pairs' must be a list of {trainer, trainee} objects")
    ew = cfg.get('extra_worked_days')
    if ew is not None and (not isinstance(ew, dict)
                           or any(not isinstance(v, list) for v in ew.values())):
        errs.append("'extra_worked_days' must be {name: [days]} (e.g. dispatch duty)")

    if errs:
        raise ScheduleConfigError(
            "Invalid config (" + config_path + "):\n  - " + "\n  - ".join(errs))
    return cfg


def _resolve_named_lists(cfg, roster, rnames):
    """Resolve exclude / exact_days / reduced_days / most_days /
    backup_eligible_extra against the roster.
    Returns (EXCLUDE set, TARGET dict, MOST set, BKX set, notes list). Fails
    loudly on unmatched/ambiguous names unless cfg['strict_names'] is False."""
    strict = cfg.get('strict_names', True)
    problems, notes = [], []

    def resolve_one(nm, where):
        h = resolve(nm, rnames)
        if len(h) == 1:
            return h[0]
        if not h:
            problems.append(f'UNMATCHED {where}: "{nm}"')
        else:
            problems.append(f'AMBIGUOUS {where}: "{nm}" -> {sorted(h)}')
        return None

    EXCLUDE = set()
    for nm in cfg.get('exclude', []):
        r = resolve_one(nm, 'exclude')
        if r:
            EXCLUDE.add(r)

    TARGET = {}
    for nm, v in cfg.get('exact_days', {}).items():
        r = resolve_one(nm, 'exact_days')
        if r:
            TARGET[r] = int(v)
    red = cfg.get('reduced_days', {})
    if red.get('names'):
        rt = int(red.get('target', 2))
        for nm in red['names']:
            r = resolve_one(nm, 'reduced_days')
            if r:
                TARGET[r] = rt

    MOST = set()
    for nm in cfg.get('most_days', []):
        r = resolve_one(nm, 'most_days')
        if r:
            MOST.add(r)

    # drivers with a day target who may ALSO take backup shifts (e.g. a
    # discipline tier held to 2 road days but allowed backup days on top)
    BKX = set()
    for nm in cfg.get('backup_eligible_extra', []):
        r = resolve_one(nm, 'backup_eligible_extra')
        if r:
            BKX.add(r)

    # backup_fallback: ORDERED groups who may take ONE backup day each when
    # the free pool can't fill a day's backups. Jose's standing order:
    # [Top performers, Solid, Underperforming/Term-review(last resort)].
    FBACK = []
    for gi, grp in enumerate(cfg.get('backup_fallback', [])):
        s = set()
        for nm in grp:
            r = resolve_one(nm, f'backup_fallback[{gi}]')
            if r:
                s.add(r)
        FBACK.append(s)

    if problems:
        msg = ("Config names did not match the roster:\n  ! "
               + "\n  ! ".join(problems)
               + "\n(Fix the names in the config, or set \"strict_names\": false "
                 "to skip them.)")
        if strict:
            raise ScheduleConfigError(msg)
        notes.extend('  ! ' + p + ' (skipped)' for p in problems)

    return EXCLUDE, TARGET, MOST, BKX, FBACK, notes


# ----------------------------------------------------------------- result ----
class Result:
    """Everything the writer / verifier / tests need from one build."""
    def __init__(self, **kw):
        self.__dict__.update(kw)


# ------------------------------------------------------------- the solver ----
def build_schedule(cfg):
    start = datetime.date.fromisoformat(cfg['start_date'])           # a Sunday
    DATEALL = {ALL[j]: start + datetime.timedelta(days=j) for j in range(7)}
    closed = set(cfg.get('closed_days', []))
    waves = cfg['waves']                                             # {day:{wavetime:count}}
    DAYS = [d for d in ALL if d in waves and d not in closed]
    routes = {d: sum(waves[d].values()) for d in DAYS}
    if cfg.get('backup_per_day'):
        backup = {d: int(cfg['backup_per_day'].get(d, 0)) for d in DAYS}
    else:
        pct = cfg.get('backup_pct', 0.15)   # standing default 15% (Jose's band: 10-20%)
        backup = {d: round(routes[d] * pct) for d in DAYS}
    # a day with no wave times can't host anything -- zero its backups too
    # (otherwise the wave-labeling step divides by an empty wave list)
    backup = {d: (backup[d] if waves[d] else 0) for d in DAYS}
    MAXC = cfg.get('max_consecutive', 5)
    PH = cfg.get('primary_hours', 10)
    BH = cfg.get('backup_hours', 2)
    FREECAP = cfg.get('free_primary_cap', 4)
    MAXPRIM = cfg.get('max_primary_days', 99)   # weekly day cap (4 = no OT). Caps ALL incl most/exact.
    HCAP = cfg.get('weekly_hours_cap', None)    # 40 = no-OT cap on ROAD days (4x10h); a fallback
                                                # 5th-day backup may sit on top (42h, Jose-approved)
    MAXTOT = cfg.get('max_total_days', 5)       # hard cap: roads + backups <= 5 days for EVERYONE
    FREETOT = cfg.get('free_total_days', 4)     # free pool (Fair): roads + backups <= 4
                                                # -> shapes are 3+1 or 2+2, never 3+2
    MAXWKND = cfg.get('max_weekend_days', None)  # cap each driver's Sat+Sun days when both open
    weekend_rule = MAXWKND is not None and WEEKEND <= set(DAYS)
    USESEED = cfg.get('use_premade_shifts', True)   # honor pre-filled shift days in avail_file
    # SOFT weekend spread (Jose 2026-07-04): when Sat+Sun both operate, aim for
    # ~1 weekend day per driver -- not half the fleet off all weekend while the
    # other half works both days. A nudge, never a rule: company need wins.
    WSPREAD = cfg.get('weekend_spread', True) and WEEKEND <= set(DAYS)

    roster = load_roster(cfg['avail_file'])
    if not USESEED:
        for dr in roster:
            dr['seed'] = set()
    rnames = [norm(d['name']) for d in roster]
    prev = load_prev_worked(cfg.get('prev_week_file'), start)
    prefs = load_prefs(cfg.get('prefs_csv'))
    merge_std = bool(cfg.get('merge_standing_unavailable'))
    for dr in roster:
        pr = prefs.get(norm(dr['name']))
        if pr:
            dr.update(usual=pr['usual'], soft=pr['soft'], present=pr['present'])
            # FEATURE: standing days-off become a hard guardrail (union with this week's)
            if merge_std:
                add = set(pr['hard']) - dr['unav']
                dr['std_added'] = add
                dr['unav'] |= add
        # prev-week worked days, matched fuzzily so a tiny name change doesn't
        # silently drop the consecutive-day carryover
        pw = prev.get(norm(dr['name']))
        if pw is None:
            hits = [k for k in prev if resolve(dr['name'], [k])]
            pw = prev[hits[0]] if len(hits) == 1 else set()
        dr['w_prev'] = set(pw)

    EXCLUDE, TARGET, MOST, BKX, FBACK, notes = _resolve_named_lists(cfg, roster, rnames)
    roster = [dr for dr in roster if norm(dr['name']) not in EXCLUDE]

    # within-tier rate ordering (Jose 2026-07-11): when two drivers sit in the
    # same priority class, the BETTER board rate (closer to 0, e.g. -14 beats
    # -24) gets more hours. Loose resolution; unknown names default to 0.
    RATE = {}
    for nm, v in cfg.get('driver_rates', {}).items():
        hits = resolve(nm, [norm(dr['name']) for dr in roster])
        if len(hits) == 1:
            RATE[hits[0]] = float(v)

    def rate_of(dr):
        return RATE.get(norm(dr['name']), 0.0)

    # the discipline tier's preferred days (Jose 2026-07-11: Sun + Sat --
    # the days nobody wants are part of the punishment)
    REDS = set()
    for nm in cfg.get('reduced_days', {}).get('names', []):
        hits = resolve(nm, [norm(dr['name']) for dr in roster])
        if len(hits) == 1:
            REDS.add(hits[0])
    REDPREF = set(cfg.get('reduced_days', {}).get('prefer_days', ['Sun', 'Sat']))

    idx = {norm(dr['name']): i for i, dr in enumerate(roster)}
    ONE = datetime.timedelta(days=1)

    # extra_worked_days (e.g. Connor's Fri/Sat dispatch duty): the driver can't
    # take a route those days (merged into unav, shown as 'Dispatch' in the
    # output) but the days COUNT as worked -- for the consecutive rule, the
    # total-days cap, and the weekend spread.
    for nm, days in cfg.get('extra_worked_days', {}).items():
        hits = [i for i, dr in enumerate(roster) if resolve(nm, [norm(dr['name'])])]
        if len(hits) != 1:
            raise ScheduleConfigError(f'extra_worked_days: "{nm}" matched {len(hits)} roster names')
        dr = roster[hits[0]]
        dr['extra'] = {d for d in days if d in ALL}
        dr['unav'] |= dr['extra']

    # meeting days block assignment the same way (but keep their own display)
    for dr in roster:
        dr['unav'] |= dr['meet']

    def worked(dr):
        return (dr['w_prev'] | {DATEALL[d] for d in dr['prim']}
                | {DATEALL[d] for d in dr['bk']} | {DATEALL[d] for d in dr['helper']}
                | {DATEALL[d] for d in dr['extra'] if d in DATEALL}
                | {DATEALL[d] for d in dr['meet'] if d in DATEALL})

    def pdays(dr):
        """Primary-day count for every cap/target: driver-of-record days PLUS
        training-helper days (a helper is out on the road all day = 10h)."""
        return len(dr['prim']) + len(dr['helper'])

    def runok(dr, dt):
        s = worked(dr) | {dt}
        n = 0; c = dt
        while c in s:
            n += 1; c -= ONE
        f = dt + ONE
        while f in s:
            n += 1; f += ONE
        return n <= MAXC

    def pcap(dr):
        n = norm(dr['name'])
        if n in TARGET:            # exact_days OR discipline: cap at the target
            base = TARGET[n]
        elif n in MOST:            # Top/Solid: maximize toward the weekly cap
            base = MAXPRIM
        else:                      # Fair: road cap; the rank enforces the soft 3-target
            base = MAXPRIM
        return min(base, MAXPRIM)

    # Day-count priority rank (Jose 2026-07-19: base floor + layered upgrades).
    # Higher = filled first / cut last. Day-independent (placement = prefsc).
    # First everyone gets a BASE (Top/Solid 3, Fair 2, discipline 1); then, while
    # routes remain, the upgrade layers in order (better rate first each layer):
    #   70  explicit exact_days (trainees, <5-routes=3, HR override) up to target
    #   -- base floor (strict tier: Top/Solid, then Fair, then discipline) --
    #   60  Top/Solid road days 1-3
    #   55  Fair road days 1-2
    #   50  discipline road day 1
    #   -- upgrades (each fully before the next) --
    #   40  Fair 2 -> 3
    #   30  discipline 1 -> 2
    #   20  Top/Solid 3 -> 4
    #   10  Fair 3 -> 4
    #    0  at cap (not a fill candidate)
    def _rank_at(dr, cur):
        n = norm(dr['name'])
        if n in TARGET and n not in REDS:      # explicit exact_days (incl. <5 routes)
            return 70 if cur < TARGET[n] else 0
        if n in REDS:                          # discipline: base 1, then upgrade to 2
            if cur < 1:
                return 50
            return 30 if cur < TARGET[n] else 0
        if n in MOST:                          # Top/Solid: base 3, then upgrade to 4
            if cur < 3:
                return 60
            return 20 if cur < MAXPRIM else 0
        # Fair: base 2, then upgrade to 3, then to 4
        if cur < 2:
            return 55
        if cur < 3:
            return 40
        return 10 if cur < MAXPRIM else 0

    def _dayscore(i, d):
        # Ordering value of giving driver i a road day on d. The greedy picks the
        # max; the rebalance moves a day to a strictly-higher-value taker.
        dr = roster[i]; n = norm(dr['name']); cur = pdays(dr)
        r = _rank_at(dr, cur)
        if r >= 60:                        # exact / Top-Solid base: fill by need
            key2 = (TARGET[n] - cur) if (n in TARGET and n not in REDS) else -cur
            return (r, key2, rate_of(dr), prefsc(dr, d))
        if r >= 50:                        # Fair/discipline base: round-robin
            return (r, -cur, rate_of(dr), prefsc(dr, d))
        return (r, rate_of(dr), -cur, prefsc(dr, d))   # upgrades: better rate first

    def wkend_ok(dr, d):
        if not weekend_rule or d not in WEEKEND:
            return True
        used = sum(1 for x in dr['prim'] + dr['bk'] + dr['helper'] if x in WEEKEND) \
            + sum(1 for x in dr['extra'] | dr['meet'] if x in WEEKEND)
        return used < MAXWKND

    FREE = lambda dr: norm(dr['name']) not in TARGET and norm(dr['name']) not in MOST

    def H(dr):
        return pdays(dr) * PH + (len(dr['bk']) + len(dr['meet'])) * BH

    def wknd_worked(dr):
        return (sum(1 for x in dr['prim'] + dr['bk'] + dr['helper'] if x in WEEKEND)
                + sum(1 for x in dr['extra'] | dr['meet'] if x in WEEKEND))

    def prefsc(dr, d):
        """Soft PLACEMENT preference of day d for a driver -- decides WHICH
        days, never how many (day counts are decided by rank/rate above this
        in the priority tuple). Ordering: seed (Jose's pre-made schedule, +40)
        > weekend spread (+/-25) > compactness (+18 adjacent / +8 near, Jose
        2026-07-11: no more Sun-Tue-Thu-Sat zigzags) > usual day (+20) >
        often-off (-12)."""
        p = 0
        if d in dr['seed']:
            p += 40
        if WSPREAD and d in WEEKEND:
            p += 25 if wknd_worked(dr) == 0 else -25
        wk = dr['prim'] + dr['helper'] + dr['bk']
        if wk:
            gap = min(abs(ALL.index(d) - ALL.index(x)) for x in wk)
            if gap == 1:
                p += 18
            elif gap == 2:
                p += 8
        if d in dr['usual']:
            p += 20
        if d in dr['soft']:
            p -= 12
        # Discipline tier works the days nobody wants (Sun/Sat) -- soft placement
        # now (Jose 2026-07-19); the day COUNT is set by the rank ladder above.
        if norm(dr['name']) in REDS:
            p += 15 if d in REDPREF else -15
        return p

    # Fill the scarcest days first: calendar order burns everyone's weekly day
    # budget on the easy early days and leaves late high-unavailability days
    # (typically Fri/Sat) unfillable. Scarcity = available drivers minus demand.
    def day_slack(d):
        have = sum(1 for dr in roster if d not in dr['unav'])
        return have - routes[d] - backup[d]
    FILL = sorted(DAYS, key=day_slack)

    # ---- PHASE 0: training pairs (Jose 2026-07-10) ----
    # A brand-new hire rides 2 BACK-TO-BACK days with the SAME trainer:
    #   day 1 -- trainer drives, new hire is the helper;
    #   day 2 -- new hire drives, trainer is the helper.
    # The pair is in ONE van, so each training day consumes exactly one route
    # slot (the driver-of-record's); the helper day counts as a worked 10h day
    # for caps/hours/consecutive but NOT toward wave counts. A trainer may
    # take a second trainee on other days (never two on the same day). Pairs
    # are placed first (most constrained) and LOCKED -- repair passes must
    # never move them.
    pslot = {d: [] for d in DAYS}
    infeasible = []
    LOCKED = set()
    PAIRLOG = []
    for pair in cfg.get('training_pairs', []):
        pt, pn = pair['trainer'], pair['trainee']
        ti = [i for i, dr in enumerate(roster) if resolve(pt, [norm(dr['name'])])]
        ni = [i for i, dr in enumerate(roster) if resolve(pn, [norm(dr['name'])])]
        if len(ti) != 1 or len(ni) != 1:
            raise ScheduleConfigError(
                f'training_pairs: "{pt}" matched {len(ti)}, "{pn}" matched {len(ni)} roster names')
        t, n = ti[0], ni[0]

        def _pair_ok(dA, dB):
            for i in (t, n):
                dr = roster[i]
                for d in (dA, dB):
                    if d in dr['unav'] or d in dr['prim'] or d in dr['helper']:
                        return False
                if pdays(dr) + 2 > pcap(dr):
                    return False
                # consecutive check with both days added
                s = worked(dr) | {DATEALL[dA], DATEALL[dB]}
                run = 0; c = DATEALL[dA]
                while c in s:
                    run += 1; c -= ONE
                f = DATEALL[dA] + ONE
                while f in s:
                    run += 1; f += ONE
                if run > MAXC:
                    return False
            return (len(pslot[dA]) < routes[dA] and len(pslot[dB]) < routes[dB])

        wins = [(ALL[j], ALL[j + 1]) for j in range(6)
                if ALL[j] in DAYS and ALL[j + 1] in DAYS]
        wins = [w for w in wins if _pair_ok(*w)]
        if not wins:
            infeasible.append(f'TRAINING: no feasible back-to-back days for '
                              f'{roster[t]["name"]} + {roster[n]["name"]}')
            continue

        # Train FIRST, solo AFTER (Jose 2026-07-10): pick the EARLIEST feasible
        # window, preferring one that leaves the trainee an available later day
        # for the solo route. Seeds/preferences do NOT delay training.
        def _solo_ok(dB):
            iB = ALL.index(dB)
            return any(ALL.index(dS) > iB and dS not in roster[n]['unav']
                       for dS in DAYS)
        dA, dB = min(wins, key=lambda w: (not _solo_ok(w[1]), ALL.index(w[0])))
        pslot[dA].append(t); roster[t]['prim'].append(dA); roster[n]['helper'].append(dA)
        pslot[dB].append(n); roster[n]['prim'].append(dB); roster[t]['helper'].append(dB)
        LOCKED.add((t, dA)); LOCKED.add((n, dB))
        roster[n]['train_done'] = ALL.index(dB)   # solo days only AFTER this
        PAIRLOG.append((roster[t]['name'], roster[n]['name'], dA, dB))

    # ---- PHASE 1: primaries (hit targets, max most, balance free by primary-day count) ----
    for d in FILL:
        dt = DATEALL[d]
        while len(pslot[d]) < routes[d]:
            cands = [i for i, dr in enumerate(roster)
                     if d not in dr['unav'] and i not in pslot[d]
                     and d not in dr['helper']
                     and (dr.get('train_done') is None or ALL.index(d) > dr['train_done'])
                     and pdays(dr) < pcap(dr) and runok(dr, dt)
                     and wkend_ok(dr, d)]
            if not cands:
                break  # PHASE 1b repairs stranded slots; shortfalls reported after it
            b = max(cands, key=lambda i: _dayscore(i, d))
            pslot[d].append(b); roster[b]['prim'].append(d)

    # ---- PHASE 1b: augmenting-path repair for slots the greedy stranded ----
    # The day-by-day greedy can leave a day short even when a full assignment
    # exists (max-flow feasible): everyone still available that day is at cap,
    # but a chain of same-count swaps frees capacity. Fill a short day either
    # directly with an under-cap driver, or by moving an at-cap driver's other
    # day here and re-filling that day recursively. Swaps keep every driver's
    # day count, so exact targets are preserved; only the terminal under-cap
    # driver (lowest-hours first) gains a day.
    def _place(i, d):
        pslot[d].append(i); roster[i]['prim'].append(d)

    def _unplace(i, d):
        pslot[d].remove(i); roster[i]['prim'].remove(d)

    def _fillable(i, d):
        dr = roster[i]
        return (d not in dr['unav'] and i not in pslot[d]
                and d not in dr['helper']
                and (dr.get('train_done') is None or ALL.index(d) > dr['train_done'])
                and runok(dr, DATEALL[d]) and wkend_ok(dr, d))

    def _augment(d, seen):
        # `seen` is one SHARED mutable set per augmentation attempt (Kuhn's
        # invariant): each day is explored at most once, so an infeasible
        # instance returns False in O(days x drivers) instead of exhausting an
        # exponential tree of copied visited-sets. A day pruned in one branch is
        # never re-tried in a sibling -- whether a day can be freed doesn't
        # depend on which driver we reached it from, so this preserves the
        # feasibility result (a path is found iff one exists).
        seen.add(d)
        under = [i for i, dr in enumerate(roster)
                 if _fillable(i, d) and pdays(dr) < pcap(dr)]
        if under:
            # lowest hours first; then better rate; then who wants the day most
            i = min(under, key=lambda i: (H(roster[i]), -rate_of(roster[i]),
                                          -prefsc(roster[i], d),
                                          norm(roster[i]['name'])))
            _place(i, d)
            return True
        # swap chains: prefer takers who want this day most, and have each
        # donor give up their least-preferred day first (never a LOCKED
        # training day)
        order = sorted((i for i, dr in enumerate(roster)
                        if _fillable(i, d) and pdays(dr) >= pcap(dr)),
                       key=lambda i: (-prefsc(roster[i], d), norm(roster[i]['name'])))
        for i in order:
            dr = roster[i]
            for e in sorted(dr['prim'], key=lambda e: (prefsc(dr, e), e)):
                if e in seen or (i, e) in LOCKED:
                    continue
                _unplace(i, e); _place(i, d)
                if _augment(e, seen):
                    return True
                _unplace(i, d); _place(i, e)
        return False

    for d in DAYS:
        while len(pslot[d]) < routes[d] and _augment(d, {d}):
            pass
        if len(pslot[d]) < routes[d]:
            infeasible.append(f'P1 INFEASIBLE {d}: filled {len(pslot[d])}/{routes[d]}')

    # ---- PHASE 1c: restore the base-floor + layer priority after repair ----
    # The scarcest-day greedy and the repair chains can leave a route slot with a
    # driver who values it LESS than someone else who could work that day (e.g. a
    # Top on a 4th day where a Fair still needs their base 2nd). Rebalance: for
    # every day, move a slot to a fillable driver whose marginal rank is strictly
    # higher than the current holder's last-held rank. Total rank-value strictly
    # increases each move, so it terminates. This enforces base-before-upgrade
    # and the layer order globally, replacing the old ad-hoc passes.
    def _move(g, f, e):
        _unplace(g, e); _place(f, e)

    def _rebalance():
        moved = True
        while moved:
            moved = False
            for d in DAYS:
                for h in list(pslot[d]):
                    if h not in pslot[d] or (h, d) in LOCKED:
                        continue
                    hval = _rank_at(roster[h], pdays(roster[h]) - 1)
                    best_t, best_key = None, None
                    for t, dt2 in enumerate(roster):
                        if t == h or not _fillable(t, d):
                            continue
                        if _rank_at(dt2, pdays(dt2)) > hval:
                            k = _dayscore(t, d)
                            if best_key is None or k > best_key:
                                best_key, best_t = k, t
                    if best_t is not None:
                        _move(h, best_t, d); moved = True
    _rebalance()

    moved = True
    while moved:
        moved = False
        frees = sorted((i for i, dr in enumerate(roster) if FREE(dr)),
                       key=lambda i: (pdays(roster[i]), -rate_of(roster[i]),
                                      norm(roster[i]['name'])))
        for f in frees:
            for g in reversed(frees):
                if pdays(roster[g]) - pdays(roster[f]) < 2:
                    continue
                cand = [e for e in roster[g]['prim']
                        if _fillable(f, e) and (g, e) not in LOCKED]
                if cand:
                    # move the day that costs the donor least and suits the
                    # receiver most (seed / weekend spread / usual)
                    e = min(cand, key=lambda e: (prefsc(roster[g], e),
                                                 -prefsc(roster[f], e), e))
                    _move(g, f, e); moved = True; break
            if moved:
                break

    # ---- PHASE 1d: target completion + Fair 3-road target ----
    # A higher-priority driver can be starved when their only eligible days are
    # saturated -- every wave is exactly full, so day-repair never fires. Lift
    # them by taking a slot from a DONOR whose held day is LESS valuable (lower
    # rank) than the day the starved driver needs (Jose 2026-07-19): e.g. a
    # discipline driver's day or a Fair's excess 4th can be pulled to complete an
    # explicit exact target or bring a Fair up to 3. Used first for exact targets,
    # then to bring every working Fair up to the 3-road target.
    def _runs_ok(dv):
        s = worked(dv)
        for dt0 in s:
            if dt0 - ONE not in s:
                run = 0; c = dt0
                while c in s:
                    run += 1; c += ONE
                if run > MAXC:
                    return False
        return True

    def _donor(v, i):
        # v may give a day to complete i iff v's LAST-held day is lower-ranked
        # (less valuable) than the day i still needs. So a discipline day or a
        # Fair's 4th can be pulled for an exact target or a Fair's 3rd, but a
        # Top/Solid day (or another driver's protected core day) never is.
        if v == i:
            return False
        dv = roster[v]
        if pdays(dv) <= 0:
            return False
        return _rank_at(dv, pdays(dv) - 1) < _rank_at(roster[i], pdays(roster[i]))

    def _complete_to(i, want):
        dr = roster[i]
        while pdays(dr) < want:
            opts = []
            for d in DAYS:
                if not _fillable(i, d):
                    continue
                vics = [v for v in pslot[d] if (v, d) not in LOCKED and _donor(v, i)]
                if vics:
                    # pull the LEAST valuable held day first (lowest rank),
                    # then busiest / most hours / worst rate
                    v = max(vics, key=lambda v: (-_rank_at(roster[v], pdays(roster[v]) - 1),
                                                 pdays(roster[v]), H(roster[v]),
                                                 -rate_of(roster[v]),
                                                 -prefsc(roster[v], d),
                                                 norm(roster[v]['name'])))
                    opts.append((prefsc(dr, d), d, v))
            if opts:
                _, d, v = max(opts, key=lambda o: (o[0], -ALL.index(o[1])))
                _unplace(v, d); _place(i, d)
                continue

            # one-level chain: no donor holds an eligible day directly -- free
            # the slot indirectly. Donor F drops day e, occupant v of needed
            # day d moves d->e, driver i takes d. Hard rules re-checked; revert
            # anything that breaks a consecutive run.
            done = False
            for d in sorted(DAYS, key=lambda x: -prefsc(dr, x)):
                if not _fillable(i, d):
                    continue
                for e in DAYS:
                    if e == d:
                        continue
                    Fs = [f for f in pslot[e] if (f, e) not in LOCKED and _donor(f, i)]
                    if not Fs:
                        continue
                    F = max(Fs, key=lambda f: (-_rank_at(roster[f], pdays(roster[f]) - 1),
                                               pdays(roster[f]), H(roster[f]),
                                               -rate_of(roster[f]),
                                               -prefsc(roster[f], e),
                                               norm(roster[f]['name'])))
                    for v in sorted(pslot[d], key=lambda x: -prefsc(roster[x], e)):
                        dv = roster[v]
                        if v in (F, i) or (v, d) in LOCKED or v in pslot[e]:
                            continue
                        if e in dv['unav'] or e in dv['helper']:
                            continue
                        if dv.get('train_done') is not None \
                                and ALL.index(e) <= dv['train_done']:
                            continue
                        if not wkend_ok(dv, e):
                            continue
                        _unplace(F, e); _unplace(v, d)
                        _place(v, e); _place(i, d)
                        if _runs_ok(dv) and _runs_ok(dr):
                            done = True; break
                        _unplace(i, d); _unplace(v, e)
                        _place(v, d); _place(F, e)
                    if done:
                        break
                if done:
                    break
            if not done:
                break   # cannot complete under the rules; verifier reports it

    # explicit exact targets only (trainee solo days, <5-routes=3, HR overrides).
    # The base floor and the Fair/discipline/Top upgrade layers are handled by
    # the rank-driven greedy + _rebalance above -- no forced completion for them.
    for i, dr in enumerate(roster):
        n = norm(dr['name'])
        if n in TARGET and n not in REDS:
            _complete_to(i, min(TARGET[n], MAXPRIM))

    # a final rebalance settles anything the exact-completion shifted.
    _rebalance()

    # ---- PHASE 2: backups -- the RATE LADDER (Jose 2026-07-20). ----
    # Backup days go to the driver with the BEST board rate who is UNDER 40
    # road+backup hours, then on down the list, one backup each, until every
    # requested slot is covered. The board rate already encodes the tier order
    # (Top ~ -1s, Solid next, Fair mid, Underperforming/Termination worst), so
    # a Top/Solid at 3 roads (30h) is served before a Fair at 30h, and the
    # discipline tier is simply the tail of the ladder -- covered LAST, but
    # covered, because the backup spots must be filled. Rate ties break by
    # tier (Top/Solid > Fair > discipline), then fewer hours, then name.
    # Rules that still bind: >=2 road days to take a backup (no backup-only
    # weeks); ONE backup each; Fair roads+backups <= free_total_days (4);
    # nobody over max_total_days (5); exact-days drivers (trainees, <5-routes,
    # benched) take no backups unless named in backup_eligible_extra.
    # OVERFLOW, only when the ladder is exhausted (nobody under 40h can take
    # an open slot): Top/Solid at 4 roads (40h) take a 5th day (42h, ~2h OT),
    # best rate first -- coverage wins as the true last resort.
    bslot = {d: [] for d in DAYS}
    BKCAP = HCAP if HCAP else 4 * PH          # the "under 40 hours" line

    def _bk_common(i, dr, d, dt):
        return (d not in dr['unav'] and d not in dr['prim'] and d not in dr['helper']
                and d not in dr['meet'] and d not in dr['extra'] and i not in bslot[d]
                and len(dr['bk']) < 1 and runok(dr, dt) and wkend_ok(dr, d)
                and pdays(dr) + len(dr['bk']) + len(dr['extra']) + len(dr['meet']) < MAXTOT)

    def _tier_rank(dr):
        n = norm(dr['name'])
        if n in MOST or n in BKX:
            return 0
        if FREE(dr):
            return 1
        return 2

    def _bk_ok(dr):
        # who is IN the rate ladder at all
        n = norm(dr['name'])
        if dr['bk'] or pdays(dr) < 2:
            return False
        if n in TARGET and n not in REDS and n not in BKX:
            return False                     # pinned exact-days: no backups
        tot = pdays(dr) + len(dr['bk']) + len(dr['extra']) + len(dr['meet'])
        if FREE(dr) and tot >= FREETOT:
            return False
        if tot >= MAXTOT:
            return False
        return H(dr) < BKCAP                 # under 40 road+backup hours

    def _bk_order(i):
        dr = roster[i]
        return (-rate_of(dr), _tier_rank(dr), H(dr), norm(dr['name']))

    def _bk_place(i, d):
        bslot[d].append(i); roster[i]['bk'].append(d)

    def _bk_augment(i, seen):
        # Serve driver i a backup: directly on their most-preferred feasible
        # day with open capacity, or by relocating a current holder of a
        # feasible-but-full day to another day they can work (the holder keeps
        # a backup -- only its day moves, so who-is-served never changes).
        # Shared `seen` set = each day explored once per attempt (Kuhn's);
        # reverts on failure.
        dr = roster[i]
        best = None
        for d in FILL:
            if d in seen or not _bk_common(i, dr, d, DATEALL[d]):
                continue
            if len(bslot[d]) < backup[d]:
                k = (prefsc(dr, d), -ALL.index(d))
                if best is None or k > best[0]:
                    best = (k, d)
        if best is not None:
            _bk_place(i, best[1])
            return True
        for d in FILL:
            if d in seen or not _bk_common(i, dr, d, DATEALL[d]):
                continue
            seen.add(d)
            for j in list(bslot[d]):
                dj = roster[j]
                bslot[d].remove(j); dj['bk'].remove(d)
                if _bk_augment(j, seen):
                    _bk_place(i, d)
                    return True
                bslot[d].append(j); dj['bk'].append(d)
        return False

    total_slots = sum(backup[d] for d in FILL)

    def _bk_filled():
        return sum(len(bslot[d]) for d in FILL)

    # Walk the rate ladder top to bottom; each driver is served via the
    # augmenting matcher, so under slot scarcity the served set is exactly the
    # rate-optimal one (a better rate never loses a slot to a worse one, and
    # nobody is stranded because a colleague took their only feasible day).
    for i in sorted((i for i, dr in enumerate(roster) if _bk_ok(dr)),
                    key=_bk_order):
        if _bk_filled() >= total_slots:
            break
        _bk_augment(i, set())

    # overflow: the ladder is spent and slots remain -> Top/Solid at 4 roads
    # take a 5th day (42h), best rate first. Coverage beats the 40h line here.
    if _bk_filled() < total_slots:
        for i in sorted((i for i, dr in enumerate(roster)
                         if norm(dr['name']) in MOST and pdays(dr) == 4
                         and not dr['bk']
                         and pdays(dr) + len(dr['extra']) + len(dr['meet']) < MAXTOT),
                        key=_bk_order):
            if _bk_filled() >= total_slots:
                break
            _bk_augment(i, set())

    # discipline-tier coverage is worth a line of its own in the report
    disc_lr = sorted((dr['name'], d) for dr in roster
                     if norm(dr['name']) in REDS and norm(dr['name']) not in BKX
                     for d in dr['bk'])
    if disc_lr:
        notes.append(f'BACKUPS: {len(disc_lr)} slot(s) covered by the '
                     f'Underperforming/Termination tier (rate-ladder tail): '
                     + ', '.join(f'{n} ({d})' for n, d in disc_lr))

    fallback_used = []   # legacy config ladder -- no longer drives the fill;
    #                      42h fifth-days are reported via the verifier instead

    # explain any backup shortfall so an under-target count never reads as a
    # silent failure -- name the open days and the reason.
    if _bk_filled() < total_slots:
        short = total_slots - _bk_filled()
        open_days = [d for d in FILL if len(bslot[d]) < backup[d]]
        notes.append(
            f'BACKUPS {_bk_filled()}/{total_slots}: {short} slot(s) open on '
            f'{"/".join(open_days)} -- nobody under {BKCAP}h with >=2 road days '
            f'can work those days, and no 4-road Top/Solid can either.')

    # ---- wave labels (primaries hit exact counts; backups spread across waves) ----
    cell = {d: {} for d in DAYS}
    for d in DAYS:
        cap = dict(waves[d]); times = list(waves[d].keys())
        for i in sorted(pslot[d], key=lambda i: roster[i]['name'].lower()):
            t = max(times, key=lambda t: cap[t]); cell[d][i] = t; cap[t] -= 1
        for k, i in enumerate(sorted(bslot[d], key=lambda i: roster[i]['name'].lower())):
            cell[d][i] = times[k % len(times)] + ' Backup'

    # training annotations: the pair shares the driver-of-record's wave; the
    # helper's cell carries the same wave + 'TRAIN helper' (NOT a route slot)
    def _short(nm):
        p = nm.split()
        return p[0] + ' ' + p[-1] if len(p) > 1 else nm

    def _annotate(d, drv, hlp):
        wave = cell[d][drv]
        cell[d][drv] = wave + ' (TRAIN drives w/ ' + _short(roster[hlp]['name']) + ')'
        cell[d][hlp] = wave + ' (TRAIN helper w/ ' + _short(roster[drv]['name']) + ')'

    nidx = {norm(dr['name']): i for i, dr in enumerate(roster)}
    for tnm, nnm, dA, dB in PAIRLOG:
        t, n = nidx[norm(tnm)], nidx[norm(nnm)]
        _annotate(dA, t, n)
        _annotate(dB, n, t)

    return Result(cfg=cfg, roster=roster, cell=cell, waves=waves, routes=routes,
                  backup=backup, DAYS=DAYS, DATEALL=DATEALL, closed=closed,
                  TARGET=TARGET, MOST=MOST, EXCLUDE=EXCLUDE, prev=prev, idx=idx,
                  MAXC=MAXC, PH=PH, BH=BH, MAXPRIM=MAXPRIM, HCAP=HCAP,
                  MAXTOT=MAXTOT, FREETOT=FREETOT, FBACK=FBACK,
                  fallback_used=fallback_used, WSPREAD=WSPREAD, USESEED=USESEED,
                  PAIRLOG=PAIRLOG, REDS=REDS, REDPREF=REDPREF, RATE=RATE,
                  MAXWKND=MAXWKND, weekend_rule=weekend_rule,
                  merge_std=merge_std, notes=notes, infeasible=infeasible)


# ------------------------------------------------------------- xlsx writer ----
def _portal_color(schedule_time):
    """Fill color for a shift cell (Jose 2026-07-10, corrected same day):
    keyed to the cell's OWN wave time -- no portal offset.
    10:05 blue / 10:25 yellow / 10:45 pink / 11:05 teal / 11:25 yellow."""
    m = re.match(r'(\d{1,2}:\d{2} [AP]M)', str(schedule_time))
    if not m:
        return None
    pal = {'10:05 AM': 'BDE7F2',   # light blue
           '10:25 AM': 'FFE380',   # yellow
           '10:45 AM': 'F9B8C6',   # pink
           '11:05 AM': '7BD5C4',   # teal ("that weird blue-green")
           '11:25 AM': 'FFE380'}   # yellow
    return pal.get(m.group(1))


def write_xlsx(res):
    cfg, roster, cell = res.cfg, res.roster, res.cell
    waves, routes, backup = res.waves, res.routes, res.backup
    DAYS, DATEALL, closed = res.DAYS, res.DATEALL, res.closed
    COLS = ALL
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = SHEET
    bold = Font(bold=True); white = Font(bold=True, color='FFFFFF')
    hf = PatternFill('solid', fgColor='305496'); bk = PatternFill('solid', fgColor='FCE4D6')
    uf = PatternFill('solid', fgColor='D9D9D9'); tf = PatternFill('solid', fgColor='E2EFDA')
    clf = PatternFill('solid', fgColor='BDD7EE')
    thin = Side('thin', color='BFBFBF'); bd = Border(thin, thin, thin, thin)
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def setc(r, c, v=None):
        cell_ = ws.cell(r, c, asciize(v))
        return cell_

    setc(1, 1, 'Time Stamp'); setc(1, 2, 'Company'); setc(1, 3, 'Station')
    for c in ('A1', 'B1', 'C1'):
        ws[c].font = bold
    setc(2, 1, fmt_timestamp(datetime.datetime.now()))
    setc(2, 2, cfg.get('company', 'JAJB LOGISTICS LLC'))
    setc(2, 3, cfg.get('station', 'WWV9'))
    for j, d in enumerate(COLS):
        c = setc(4, 3 + j, f"{d}, {DATEALL[d].strftime('%d/%b')}")
        c.font = white; c.fill = hf; c.alignment = ctr; c.border = bd
    for j, h in enumerate(['Associate Name', 'Transporter ID'], 1):
        c = setc(4, j, h); c.font = white; c.fill = hf; c.alignment = ctr; c.border = bd
    setc(5, 1, 'Total Scheduled (routes + backup)').font = bold
    ws.cell(5, 1).fill = tf; ws.cell(5, 2).fill = tf
    for j, d in enumerate(COLS):
        if d in closed or d not in DAYS:
            txt = 'CLOSED'
        else:
            txt = f'{routes[d]}+{backup[d]}={routes[d] + backup[d]}'
        c = setc(5, 3 + j, txt); c.font = bold; c.fill = tf; c.alignment = ctr; c.border = bd
    order = sorted(range(len(roster)), key=lambda i: roster[i]['name'].lower()); r = 6
    for i in order:
        dr = roster[i]
        setc(r, 1, dr['name']).border = bd
        setc(r, 2, dr['tid']).border = bd
        for j, d in enumerate(COLS):
            c = ws.cell(r, 3 + j); c.alignment = ctr; c.border = bd
            if d in closed or d not in DAYS:
                c.fill = clf; continue
            v = cell[d].get(i)
            if v:
                c.value = asciize(v)
                wavecol = _portal_color(v)
                if 'Backup' in v:
                    c.fill = bk
                elif 'TRAIN helper' in v:
                    # portal legend: Helper - orange
                    c.fill = PatternFill('solid', fgColor='F5A26B')
                elif wavecol:
                    c.fill = PatternFill('solid', fgColor=wavecol)
            elif d in dr['meet']:
                # do-not-touch: the pre-entered meeting cell, preserved
                # verbatim; portal legend: Meeting - lavender
                c.value = asciize(dr['meet_txt'][d])
                c.fill = PatternFill('solid', fgColor='E5C5EA')
                c.font = Font(italic=True, bold=True)
            elif d in dr['extra']:
                # portal legend: Dispatcher - yellow-green
                c.value = 'Dispatch'
                c.fill = PatternFill('solid', fgColor='C9E265')
                c.font = Font(italic=True)
            elif d in dr['unav']:
                c.value = 'Unavailable'; c.fill = uf
                c.font = Font(italic=True, color='808080')
        r += 1
    ws.freeze_panes = 'C6'
    ws.column_dimensions['A'].width = 24; ws.column_dimensions['B'].width = 17
    # Jose 2026-07-10 (corrected): normal column WIDTH, row HEIGHT 40px
    # (= 30 pt; 1 pt = 4/3 px).
    for col in 'CDEFGHI':
        ws.column_dimensions[col].width = 15
    for rr_ in range(6, r):
        ws.row_dimensions[rr_].height = 30

    # By Day
    ws2 = wb.create_sheet('By Day'); ws2.column_dimensions['A'].width = 14
    for col in 'BCDEFGH':
        ws2.column_dimensions[col].width = 24
    ws2.cell(1, 1, asciize(f"{cfg.get('week_label', 'Week')} - {cfg.get('station', 'WWV9')}")).font = Font(bold=True, size=13)
    for j, d in enumerate(DAYS):
        c = ws2.cell(3, 2 + j, f"{d} {DATEALL[d].strftime('%d/%b')}")
        c.font = white; c.fill = hf; c.alignment = ctr; c.border = bd
    alltimes = sorted({t for d in DAYS for t in waves[d]})
    rr = 4
    for t in alltimes:
        ws2.cell(rr, 1, asciize(t)).font = bold
        for j, d in enumerate(DAYS):
            ws2.cell(rr, 2 + j, waves[d].get(t, '')).alignment = ctr
        rr += 1
    for lab, val in [('Routes', routes), ('Backup', backup)]:
        ws2.cell(rr, 1, lab).font = bold
        for j, d in enumerate(DAYS):
            ws2.cell(rr, 2 + j, val[d]).alignment = ctr
        rr += 1
    ws2.cell(rr, 1, 'Total').font = bold
    for j, d in enumerate(DAYS):
        ws2.cell(rr, 2 + j, routes[d] + backup[d]).font = bold
    start = rr + 2
    ws2.cell(start - 1, 1, 'Roster by wave (Backups shaded)').font = Font(bold=True, italic=True)
    for j, d in enumerate(DAYS):
        lst = []
        for t in alltimes:
            names = sorted(
                roster[i]['name'] + (' [TRAIN drives]' if 'TRAIN drives' in v
                                     else ' [TRAIN helper]' if 'TRAIN helper' in v else '')
                for i, v in cell[d].items()
                if v.startswith(t) and 'Backup' not in v)
            if names:
                lst.append((f'- {t} -', 'h')); lst += [(n, 'p') for n in names]
        bn = sorted(roster[i]['name'] for i, v in cell[d].items() if 'Backup' in v)
        lst.append(('- Backup -', 'h')); lst += [(n, 'b') for n in bn]
        for k, (n, ty) in enumerate(lst):
            c = ws2.cell(start + k, 2 + j, asciize(n)); c.border = bd
            if ty == 'h':
                c.font = Font(bold=True, color='305496')
            elif ty == 'b':
                c.fill = bk
    ws2.freeze_panes = 'B4'

    # Training -- one row per trainer/trainee pair this week. Day 1 is the day
    # the trainer drives (trainee rides along); Day 2 is the day the trainee
    # drives solo with the trainer supporting. Source: res.PAIRLOG.
    ws3 = wb.create_sheet('Training')
    ws3.cell(1, 1, asciize(
        f"{cfg.get('week_label', 'Week')} - {cfg.get('station', 'WWV9')} - Training pairs"
    )).font = Font(bold=True, size=13)
    heads = ['Trainer', 'Trainer ID', 'Trainee', 'Trainee ID',
             'Day 1 - trainer drives', 'Day 2 - trainee drives']
    for j, h in enumerate(heads, 1):
        c = ws3.cell(3, j, asciize(h))
        c.font = white; c.fill = hf; c.alignment = ctr; c.border = bd
    tid_of = {dr['name']: dr['tid'] for dr in roster}
    if res.PAIRLOG:
        for k, (tnm, nnm, dA, dB) in enumerate(sorted(res.PAIRLOG)):
            vals = [tnm, tid_of.get(tnm, ''), nnm, tid_of.get(nnm, ''),
                    f"{dA} {DATEALL[dA].strftime('%d/%b')}" if dA in DATEALL else dA,
                    f"{dB} {DATEALL[dB].strftime('%d/%b')}" if dB in DATEALL else dB]
            for j, v in enumerate(vals, 1):
                c = ws3.cell(4 + k, j, asciize(v)); c.border = bd
                if j >= 5:
                    c.alignment = ctr
    else:
        ws3.cell(4, 1, 'No training pairs scheduled this week.').font = \
            Font(italic=True, color='808080')
    for j, w in enumerate([24, 17, 24, 17, 22, 22]):
        ws3.column_dimensions[chr(ord('A') + j)].width = w
    ws3.freeze_panes = 'A4'

    wb.save(cfg['out'])
    print('Saved', cfg['out'])


# --------------------------------------------------------------- verifier ----
def check_invariants(res):
    """Compute every invariant independently from the Result. Returns a dict of
    structured results (used by both the CLI summary and the test suite)."""
    roster, cell = res.roster, res.cell
    waves, routes, backup = res.waves, res.routes, res.backup
    DAYS, DATEALL = res.DAYS, res.DATEALL
    TARGET, MOST = res.TARGET, res.MOST
    MAXC, PH, BH = res.MAXC, res.PH, res.BH
    ONE = datetime.timedelta(days=1)
    errs = []

    def pdy(dr):
        # primary-day count for caps/targets: driver-of-record + training-helper days
        return len(dr['prim']) + len(dr['helper'])

    def wavekey(v):
        m = re.match(r'(\d{1,2}:\d{2} [AP]M)', v)
        return m.group(1) if m else v

    def is_route(v):
        # a route slot: not a backup, not a training HELPER (the helper rides
        # in the driver-of-record's van -- one route per training pair per day)
        return 'Backup' not in v and 'TRAIN helper' not in v

    # inv 3: exact per-wave route counts; inv 1: no one on an Unavailable day
    for d in DAYS:
        cnt = Counter(wavekey(v) for v in cell[d].values() if is_route(v))
        for t, n in waves[d].items():
            if cnt.get(t, 0) != n:
                errs.append(f'WAVE {d} {t}: {cnt.get(t, 0)} != {n}')
        for i in cell[d]:
            if d in roster[i]['unav'] and d not in roster[i]['extra'] \
                    and d not in roster[i]['meet']:
                errs.append(f'UNAVAIL violated: {roster[i]["name"]} {d}')
            if d in roster[i]['meet']:
                errs.append(f'MEETING-DAY touched: {roster[i]["name"]} {d} '
                            f'(got {cell[d][i]!r} on a do-not-touch meeting day)')

    # inv 2: max consecutive worked days (incl prev-week tail, helper days,
    # and extra worked days like dispatch duty)
    mx = 0
    for dr in roster:
        wd = (set(dr['w_prev']) | {DATEALL[d] for d in dr['prim']}
              | {DATEALL[d] for d in dr['bk']} | {DATEALL[d] for d in dr['helper']}
              | {DATEALL[d] for d in dr['extra'] if d in DATEALL}
              | {DATEALL[d] for d in dr['meet'] if d in DATEALL})
        for dt in wd:
            if dt - ONE not in wd:
                n = 0; c = dt
                while c in wd:
                    n += 1; c += ONE
                mx = max(mx, n)
                if n > MAXC:
                    errs.append(f'CONSEC>{MAXC}: {dr["name"]} run={n}')
        overlaps = (set(dr['prim']) & set(dr['bk'])) | (set(dr['prim']) & set(dr['helper'])) \
            | (set(dr['bk']) & set(dr['helper']))
        if overlaps:
            errs.append(f'DUP-DAY: {dr["name"]} (two roles on {sorted(overlaps)})')

    def H(dr):
        return pdy(dr) * PH + len(dr['bk']) * BH

    # inv 5: targets. Explicit exact_days must be hit unless the driver's own
    # availability blocks it (undershoot is a NOTE, not a build error). Discipline
    # and Fair are SOFT tier targets allocated by the priority ladder and are
    # never a hard error (Jose 2026-07-19 tier rework).
    capd = res.MAXPRIM
    REDS = res.REDS
    target_bad = []
    target_short = []
    for n in TARGET:
        if n not in res.idx:
            continue
        want = min(TARGET[n], capd)
        got = pdy(roster[res.idx[n]])
        if n in REDS:
            continue                       # discipline: soft, day count informational
        if got > want:                     # over target / a benched driver got scheduled
            target_bad.append((n, want, got))
            errs.append(f'TARGET {n}: want {want} got {got} (over)')
        elif got < want:                   # availability-limited: a note, not an error
            target_short.append((n, want, got))

    # inv 4: no road-day OT, backups are top-ups, no backup-only weeks.
    # HCAP caps ROAD hours (4x10h=40). A fallback 5th-day backup may sit on
    # top (42h total, Jose-approved 2026-07-04) -- reported, not an error.
    HCAP = res.HCAP
    over_cap = [dr['name'] for dr in roster if HCAP and pdy(dr) * PH > HCAP]
    for nm in over_cap:
        errs.append(f'OT: {nm} road days over {HCAP}h')
    fifth_day = [(dr['name'], H(dr)) for dr in roster
                 if HCAP and pdy(dr) * PH <= HCAP and H(dr) > HCAP]
    over_days = [dr['name'] for dr in roster if pdy(dr) > capd]
    for nm in over_days:
        errs.append(f'DAYCAP: {nm} over {capd} primary days')
    # total worked-days caps: 5 for everyone (incl. dispatch/extra days),
    # 4 for the free pool (Fair shapes)
    MAXTOT, FREETOT = res.MAXTOT, res.FREETOT
    over_tot = [dr['name'] for dr in roster
                if pdy(dr) + len(dr['bk']) + len(dr['extra']) + len(dr['meet']) > MAXTOT]
    for nm in over_tot:
        errs.append(f'TOTDAYS: {nm} over {MAXTOT} worked days')
    free_shape_bad = [dr['name'] for dr in roster
                      if FREE_name(dr, TARGET, MOST)
                      and pdy(dr) + len(dr['bk']) > FREETOT]
    for nm in free_shape_bad:
        errs.append(f'FAIR-SHAPE: {nm} roads+backups over {FREETOT}')
    backup_only = [dr['name'] for dr in roster if dr['bk'] and not pdy(dr)]
    for nm in backup_only:
        errs.append(f'BACKUP-ONLY: {nm}')
    backup_under2 = [dr['name'] for dr in roster if dr['bk'] and pdy(dr) < 2]
    for nm in backup_under2:
        errs.append(f'BACKUP<2PRIMARY: {nm}')

    # Fair backup floor (revised 2026-07-16): a 2-road free-pool driver should
    # carry at least 1 backup (22h). Their 2nd backup is now a bonus that ranks
    # BELOW a 3-day Top/Solid's first backup, so <2 is expected, not a warning;
    # only 0 backups (a bare 20h week) is flagged. Availability / supply can
    # legitimately block even that, so it's a WARNING, not a hard error.
    # (Fair no longer take backups under the base-floor model, so a Fair at 2
    # road with no backup is expected -- nothing to flag.)
    floor_unmet = []

    free = [dr for dr in roster if FREE_name(dr, TARGET, MOST)]
    fh = sorted(H(dr) for dr in free)
    pool = dict(min=min(fh) if fh else 0, max=max(fh) if fh else 0,
                avg=round(sum(fh) / len(fh), 1) if fh else 0,
                dist=dict(sorted(Counter(fh).items())))
    per_day = {d: dict(routes=sum(1 for v in cell[d].values() if is_route(v)),
                       backup=sum(1 for v in cell[d].values() if 'Backup' in v))
               for d in DAYS}

    # usual-day adherence: week-to-week stickiness (drivers with history only)
    tot_u = on_u = 0
    off_usual = []
    for dr in roster:
        if not dr['usual']:
            continue
        for d in dr['prim']:
            tot_u += 1
            if d in dr['usual']:
                on_u += 1
            else:
                off_usual.append((dr['name'], d))
    usual_pct = round(100 * on_u / tot_u, 1) if tot_u else None

    # seed adherence: how much of Jose's pre-made schedule survived. A seed
    # day counts as kept if the driver works it (road or backup). Dropped
    # seeds are expected when a tier cap / wave count / rule forced it.
    seed_tot = seed_kept = 0
    seed_dropped = []
    for dr in roster:
        wk = set(dr['prim']) | set(dr['bk']) | set(dr['helper'])
        for d in dr['seed']:
            if d not in res.DAYS:      # seed on a closed/non-operating day
                continue
            seed_tot += 1
            if d in wk:
                seed_kept += 1
            else:
                seed_dropped.append((dr['name'], d))
    seed_pct = round(100 * seed_kept / seed_tot, 1) if seed_tot else None

    # weekend spread: distribution of worked weekend days (Sat+Sun both open)
    wknd_dist = None
    if res.WSPREAD:
        wd = Counter(sum(1 for x in dr['prim'] + dr['bk'] + dr['helper'] if x in WEEKEND)
                     + sum(1 for x in dr['extra'] if x in WEEKEND)
                     for dr in roster
                     if dr['prim'] or dr['bk'] or dr['helper'] or dr['extra'])
        wknd_dist = dict(sorted(wd.items()))

    meetings = [(dr['name'], d) for dr in roster for d in sorted(dr['meet'])]

    # discipline-tier placement (Jose 2026-07-11: their days should be the
    # preferred punishment days -- default Sun/Sat -- whenever possible)
    red_place = None
    if res.REDS:
        onpref = offpref = 0
        for dr in roster:
            if norm(dr['name']) in res.REDS:
                for d in dr['prim']:
                    if d in res.REDPREF:
                        onpref += 1
                    else:
                        offpref += 1
        red_place = dict(on_preferred=onpref, weekday=offpref)

    # Fair floor (Jose 2026-07-11): every available free-pool driver should
    # reach 2 road days before any 4th day is granted
    floor2_road = [(dr['name'], pdy(dr)) for dr in roster
                   if FREE_name(dr, TARGET, MOST) and 0 < pdy(dr) < 2]

    return dict(errors=errs, max_consec=mx, target_bad=target_bad,
                target_short=target_short,
                over_cap=over_cap, over_days=over_days, backup_only=backup_only,
                backup_under2=backup_under2, floor_unmet=floor_unmet,
                fifth_day=fifth_day, over_tot=over_tot,
                free_shape_bad=free_shape_bad,
                pool=pool, per_day=per_day, meetings=meetings,
                red_place=red_place, floor2_road=floor2_road,
                usual_pct=usual_pct, usual_n=tot_u, off_usual=off_usual,
                seed_pct=seed_pct, seed_n=seed_tot, seed_dropped=seed_dropped,
                wknd_dist=wknd_dist)


def FREE_name(dr, TARGET, MOST):
    n = norm(dr['name'])
    return n not in TARGET and n not in MOST


def print_summary(res, chk):
    print('\n=== VERIFICATION ===')
    for d in res.DAYS:
        pd = chk['per_day'][d]
        pct = round(100 * pd['backup'] / res.routes[d], 1) if res.routes[d] else 0
        print(f"  {d}: routes {pd['routes']}/{res.routes[d]}  "
              f"backup {pd['backup']}/{res.backup[d]} ({pct}%)")
    if res.PAIRLOG:
        print('  training pairs (day1 trainer drives / day2 trainee drives):')
        for tnm, nnm, dA, dB in res.PAIRLOG:
            print(f'    {tnm} + {nnm}: {dA} -> {dB}')
    if chk.get('meetings'):
        print('  meeting days preserved (do-not-touch):',
              '; '.join(f'{n} ({d})' for n, d in chk['meetings']))
    print('  exact targets all met (capped):', not chk['target_bad'],
          ('' if not chk['target_bad'] else chk['target_bad']))
    if chk.get('target_short'):
        print('  exact targets short (availability-limited, not an error):',
              chk['target_short'])
    p = chk['pool']
    print(f"  Fair-driver hours: min {p['min']} max {p['max']} avg {p['avg']} dist {p['dist']}")
    print('  road days over', res.HCAP, 'h (OT):', chk['over_cap'] or 'none')
    print('  backup-only weeks:', chk['backup_only'] or 'none')
    print('  backups under 2 primary:', chk['backup_under2'] or 'none')
    print('  working Fair drivers under the 2-road base floor:',
          chk.get('floor2_road') or 'none')
    if chk.get('red_place') is not None:
        rp = chk['red_place']
        print(f"  discipline-tier road days: {rp['on_preferred']} on "
              f"{'/'.join(sorted(res.REDPREF))}, {rp['weekday']} elsewhere")
    print('  5th-day fallback backups (42h, review):',
          chk.get('fifth_day') or 'none')
    if res.fallback_used:
        by_grp = {}
        for gi, nm, d in res.fallback_used:
            by_grp.setdefault(gi, []).append(f'{nm} ({d})')
        for gi in sorted(by_grp):
            print(f'  fallback group {gi + 1} used:', '; '.join(by_grp[gi]))
    if res.merge_std:
        added = [(dr['name'], sorted(dr['std_added'])) for dr in res.roster if dr['std_added']]
        print('  standing days-off merged:', added or 'none added (all already submitted)')
    if res.weekend_rule:
        print('  weekend cap:', res.MAXWKND, 'day(s) per driver (Sat+Sun both open)')
    if chk['usual_pct'] is not None:
        print(f"  usual-day adherence: {chk['usual_pct']}% of {chk['usual_n']} "
              f"primary shifts on the driver's usual day")
    if chk.get('seed_pct') is not None:
        print(f"  pre-made schedule kept: {chk['seed_pct']}% of {chk['seed_n']} "
              f"pre-entered shift days")
        if chk['seed_dropped']:
            print('    dropped (rule/cap forced):',
                  '; '.join(f'{n} ({d})' for n, d in chk['seed_dropped'][:25]),
                  ('... +%d more' % (len(chk['seed_dropped']) - 25)
                   if len(chk['seed_dropped']) > 25 else ''))
    if chk.get('wknd_dist') is not None:
        print('  weekend days per working driver (soft target = 1):',
              chk['wknd_dist'])
    print('  max consecutive run:', chk['max_consec'], '(cap', res.MAXC, ')')
    for line in res.infeasible:
        print('  !', line)
    for line in res.notes:
        print(line)
    print('  ERRORS:', len(chk['errors']))
    for e in chk['errors'][:20]:
        print('     ', e)


# -------------------------------------------------------------------- main ----
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--config', required=True)
    a = ap.parse_args()
    try:
        cfg = load_config(a.config)
    except ScheduleConfigError as e:
        print('CONFIG ERROR:\n' + str(e), file=sys.stderr)
        sys.exit(2)
    try:
        res = build_schedule(cfg)
    except ScheduleConfigError as e:
        print('BUILD ERROR:\n' + str(e), file=sys.stderr)
        sys.exit(2)
    write_xlsx(res)
    chk = check_invariants(res)
    print_summary(res, chk)
    if chk['errors'] or res.infeasible:
        sys.exit(1)


if __name__ == "__main__":
    main()
